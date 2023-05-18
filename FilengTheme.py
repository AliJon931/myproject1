from openpyxl import Workbook,load_workbook
import datetime

wb = Workbook()
wl = load_workbook("Baza.xlsx")
wk = wl.active
ws = wb.active
Data = []
Data_viwe = []




text = """
1 - Maxsulot qoshish
2 - Maxsulot korish
3 - Maxsulot sotish
4 - Xisobot 
5 - Muddat
"""

while True:
    print(text)
    a = input("Bolimini tanlang >> ")
    if a == "1" :
        print("Qoshish")

        for i in range(1, wk.max_row+1):
            A = f'A{i}'
            B = f'B{i}'
            C = f'C{i}'
            D = f'D{i}'
            E = f'E{i}'
            F = f'F{i}'
            Data.append([wk[A].value, wk[B].value,wk[C].value,wk[D].value,wk[E].value,wk[F].value])

        while True:
            nomi = input("Nomi : ")
            if nomi == '0':
                break
            narx = int(input("Narxi : "))
            soni = int(input("Soni : "))
            foiz = (narx / 100) * 120
            code_shtrix = int(input("Shtrix Kodi : "))
            now = datetime.datetime.now().strftime('%d/%m/%y')
            Data.append([nomi,narx,soni,foiz,code_shtrix,now])

        for i in Data:
            ws.append(i)
        wb.save("Baza.xlsx")    

    elif a == "2":
        print("Maxsulotlar\n")

        for i in range(1, wk.max_row+1):
            A = f'A{i}'
            B = f'B{i}'
            C = f'C{i}'
            D = f'D{i}'
            E = f'E{i}'
            F = f'F{i}'
            Data_viwe.append([wk[A].value, wk[B].value,wk[C].value,wk[D].value,wk[E].value,wk[F].value])

        for i in Data_viwe:
            print(i)


    elif a == '3':
        print("Sotish")
        maxnom = input('Maxsulot nomi : ')
        # maxson = int(input("sotilgan Soni : "))
        sotish = []
        sotish2=[]
        # n=-1
        # maxnom = input('Maxsulot nomi : ')
        # for i in range(2, wk.max_row+1):
        #     A=f"A{i}"
        #     sotish.append(wk[A].value)
        # for j in sotish:
        #     if maxnom == j:
        #         n = sotish.index(j)
        # if n == -1:
        #     print("Mahsulot Topilmadi! ")
        #     continue
        # for i in range(2, wk.max_row + 1):
        #     B = f'B{i}'
        for i in range(1, wk.max_row+1):
            A = f'A{i}'
            B = f'B{i}'
            C = f'C{i}'
            D = f'D{i}'
            E = f'E{i}'
            F = f'F{i}'
            sotish.append([wk[A].value, wk[B].value,wk[C].value,wk[D].value,wk[E].value,wk[F].value])

            # temp = sotish.index(maxnom)
            for i in sotish:
                if i == maxnom:
                    maxson = int(input('Maxsulot soni :'))
                    if i[1] >= maxson:
                        print("Maxsulot sotildi!")
                        sotish2.append(i[0])
                        sotish2.append(i[1] - maxson)
                        continue

                sotish2.append(i)
                      


            # if sotish[temp + 2] >= maxson:
            #     sotish[temp + 2] -= maxson
                for i in sotish2:
                    ws.append(i)
                    wb.save("Baza.xlsx") 
            # else :
            #     print("Hatolik!")
            

    elif a == '4':
        print('Xisobot') 
        
    elif a == 5 :
        sana1=[]
        for i in range(2, wk.max_row + 1):
            G = f"G{i}"
            sana1.append(wk[G].value)
        for i in sana1:
            dd = i.split('/')
            x = datetime.datetime(int(dd[0]),int(dd[1]),int(dd[2]))
            print((x - datetime.datetime.now()).days)

        print("Muddati")
 